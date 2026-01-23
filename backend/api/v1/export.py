"""导出相关API"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from datetime import datetime
from typing import Optional
from database.models import User
from database.operations import DatabaseManager
from api.deps import get_db, get_current_user

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from urllib.parse import quote

router = APIRouter(prefix="/export", tags=["导出"])


def get_excel_output_path():
    """获取Excel文件输出路径"""
    output_dir = "exports"
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def create_border(border_style="thin", color="000000"):
    """创建边框样式"""
    return Border(
        left=Side(style=border_style, color=color),
        right=Side(style=border_style, color=color),
        top=Side(style=border_style, color=color),
        bottom=Side(style=border_style, color=color)
    )


def create_excel(accounts, ledger_name):
    """创建Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "账单明细"

    # 定义样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_border = create_border("thin", "D9D9D9")

    cell_font = Font(name='微软雅黑', size=11)
    cell_alignment = Alignment(horizontal="left", vertical="center")
    cell_border = create_border("thin", "D9D9D9")

    # 设置列宽
    column_widths = {
        'A': 15,  # 日期
        'B': 10,  # 类型
        'C': 15,  # 分类
        'D': 25,  # 商品名称
        'E': 15,  # 金额
        'F': 20,  # 地点
        'G': 30,  # 备注
        'H': 40   # 附件
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 表头
    headers = ['日期', '类型', '分类', '商品名称', '金额', '地点', '备注', '附件']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # 数据行
    for row_idx, account in enumerate(accounts, 2):
        # 日期
        cell = ws.cell(row=row_idx, column=1)
        cell.value = account.get('transaction_date', '')
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = cell_border

        # 类型
        cell = ws.cell(row=row_idx, column=2)
        cell.value = account.get('transaction_type', '')
        cell.font = cell_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

        # 分类
        cell = ws.cell(row=row_idx, column=3)
        cell.value = account.get('category', '未分类')
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = cell_border

        # 商品名称
        cell = ws.cell(row=row_idx, column=4)
        cell.value = account.get('item_name', '')
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = cell_border

        # 金额
        cell = ws.cell(row=row_idx, column=5)
        amount = account.get('amount', 0)
        cell.value = f"¥{amount:.2f}" if amount else ''
        cell.font = cell_font
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = cell_border

        # 地点
        cell = ws.cell(row=row_idx, column=6)
        cell.value = account.get('merchant_name', '')
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = cell_border

        # 备注
        cell = ws.cell(row=row_idx, column=7)
        cell.value = account.get('notes', '')
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = cell_border

        # 附件
        cell = ws.cell(row=row_idx, column=8)
        image_path = account.get('image_path', '')
        if image_path:
            # 如果是相对路径，转换为完整URL
            if not image_path.startswith('http'):
                image_path = f"/static/{image_path}"
        cell.value = image_path
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = cell_border

    # 冻结第一行
    ws.freeze_panes = "A2"

    # 设置行高
    ws.row_dimensions[1].height = 25

    return wb


@router.get("/accounts/{ledger_id}")
async def export_accounts(
    ledger_id: int,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: DatabaseManager = Depends(get_db)
):
    """导出账本账单为Excel"""
    # 验证账本所有权
    ledger = db.get_ledger_by_id(ledger_id)
    if not ledger or ledger.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="账本不存在")

    # 获取账本名称
    ledger_name = ledger.name

    # 获取账单数据
    accounts = db.get_ledger_accounts(ledger_id)

    # 过滤日期范围
    if start_date:
        accounts = [a for a in accounts if a.get('transaction_date', '') >= start_date]
    if end_date:
        accounts = [a for a in accounts if a.get('transaction_date', '') <= end_date]

    # 按日期降序排序
    accounts.sort(key=lambda x: x.get('transaction_date', ''), reverse=True)

    # 创建Excel
    wb = create_excel(accounts, ledger_name)

    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{ledger_name}_账单明细_{timestamp}.xlsx"
    filepath = os.path.join(get_excel_output_path(), filename)

    # 保存文件
    wb.save(filepath)

    # 返回文件，使用 RFC 2231 编码中文文件名
    encoded_filename = quote(filename, safe='')
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )
